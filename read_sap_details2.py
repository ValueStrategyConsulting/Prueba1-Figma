import openpyxl
import os

base = r"c:\Users\Usuario\Desktop\ASSET-MANAGEMENT-SOFTWARE\MAINTENANCE STRATEGY DEVELOPMENT METHODOLOGY AND EXISTING LEGACY SOFTWARE\sap-upload-sheets-template-(MIS-P01-BF01-F740-7310)"

files = ["Maintenance Item.xlsx", "Task List.xlsx", "Work Plan.xlsx"]

for fname in files:
    path = os.path.join(base, fname)
    print("=" * 120)
    print(f"FILE: {fname} - DETAILED ANALYSIS")
    print("=" * 120)
    wb = openpyxl.load_workbook(path)
    for sname in wb.sheetnames:
        ws = wb[sname]

        # Data validations
        print(f"\n  DATA VALIDATIONS:")
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                print(f"    Range: {dv.sqref}, Type: {dv.type}, Formula1: {dv.formula1}, Formula2: {dv.formula2}")
        else:
            print(f"    None found")

        # Data row cell details with types
        print(f"\n  DATA ROW(S) CELL DETAILS:")
        for row_idx in range(2, min(ws.max_row + 1, 10)):
            print(f"    --- Row {row_idx} ---")
            for cell in ws[row_idx]:
                if cell.value is not None:
                    print(f"    {cell.coordinate} = {repr(cell.value)} (type={type(cell.value).__name__}, numfmt={cell.number_format})")

        # Comments
        print(f"\n  CELL COMMENTS:")
        found_comment = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.comment:
                    print(f"    {cell.coordinate}: {cell.comment.text} (by {cell.comment.author})")
                    found_comment = True
        if not found_comment:
            print(f"    None found")

        # Conditional formatting
        print(f"\n  CONDITIONAL FORMATTING:")
        if ws.conditional_formatting:
            for cf in ws.conditional_formatting:
                print(f"    Range: {cf}, Rules: {len(cf.rules)}")
        else:
            print(f"    None found")

    print()
    wb.close()
