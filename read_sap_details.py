import openpyxl
import os

base = r"c:\Users\Usuario\Desktop\ASSET-MANAGEMENT-SOFTWARE\MAINTENANCE STRATEGY DEVELOPMENT METHODOLOGY AND EXISTING LEGACY SOFTWARE\sap-upload-sheets-template-(MIS-P01-BF01-F740-7310)"

files = ["Maintenance Item.xlsx", "Task List.xlsx", "Work Plan.xlsx"]

for fname in files:
    path = os.path.join(base, fname)
    print("=" * 120)
    print(f"FILE: {fname} - DETAILED ANALYSIS")
    print("=" * 120)
    wb = openpyxl.load_workbook(path)
    for sname in wb.sheetnames:
        ws = wb[sname]

        # Check for data validations
        print(f"\n  DATA VALIDATIONS:")
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                print(f"    Range: {dv.sqref}, Type: {dv.type}, Formula1: {dv.formula1}, Formula2: {dv.formula2}, Allow blank: {dv.allow_blank}, Operator: {dv.operator}")
        else:
            print(f"    None found")

        # Check column widths
        print(f"\n  COLUMN WIDTHS:")
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                print(f"    Column {col_letter}: width={dim.width}")

        # Cell formatting details for header row
        print(f"\n  HEADER ROW CELL DETAILS:")
        for cell in ws[1]:
            if cell.value is not None:
                font = cell.font
                fill = cell.fill
                alignment = cell.alignment
                border = cell.border
                print(f"    {cell.coordinate} = {repr(cell.value)}")
                print(f"      Font: bold={font.bold}, italic={font.italic}, size={font.size}, color={font.color}")
                print(f"      Fill: fgColor={fill.fgColor}, bgColor={fill.bgColor}, patternType={fill.patternType}")
                print(f"      NumFmt: {cell.number_format}")

        # Data row cell details
        print(f"\n  DATA ROW(S) CELL DETAILS:")
        for row_idx in range(2, min(ws.max_row + 1, 10)):
            print(f"    --- Row {row_idx} ---")
            for cell in ws[row_idx]:
                if cell.value is not None:
                    print(f"    {cell.coordinate} = {repr(cell.value)} (type={type(cell.value).__name__}, numfmt={cell.number_format})")

        # Check for named ranges
        print(f"\n  DEFINED NAMES (workbook level):")
        for dn in wb.defined_names.definedName:
            print(f"    {dn.name} = {dn.attr_text}")

        # Check for comments
        print(f"\n  CELL COMMENTS:")
        found_comment = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.comment:
                    print(f"    {cell.coordinate}: {cell.comment.text} (by {cell.comment.author})")
                    found_comment = True
        if not found_comment:
            print(f"    None found")

        # Check conditional formatting
        print(f"\n  CONDITIONAL FORMATTING:")
        if ws.conditional_formatting:
            for cf in ws.conditional_formatting:
                print(f"    Range: {cf}, Rules: {len(cf.rules)}")
                for rule in cf.rules:
                    print(f"      Type: {rule.type}, Formula: {rule.formula}, Priority: {rule.priority}")
        else:
            print(f"    None found")

    print()
    wb.close()
