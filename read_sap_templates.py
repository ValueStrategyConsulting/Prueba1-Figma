import openpyxl
import os

base = r"c:\Users\Usuario\Desktop\ASSET-MANAGEMENT-SOFTWARE\MAINTENANCE STRATEGY DEVELOPMENT METHODOLOGY AND EXISTING LEGACY SOFTWARE\sap-upload-sheets-template-(MIS-P01-BF01-F740-7310)"

files = ["Maintenance Item.xlsx", "Task List.xlsx", "Work Plan.xlsx"]

for fname in files:
    path = os.path.join(base, fname)
    print("=" * 120)
    print(f"FILE: {fname}")
    print("=" * 120)
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n--- Sheet: \"{sname}\" (rows={ws.max_row}, cols={ws.max_column}) ---")

        # Print merged cells info
        if ws.merged_cells.ranges:
            print(f"  Merged cells: {[str(m) for m in ws.merged_cells.ranges]}")

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=False), start=1):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    # Also get cell formatting info
                    fmt = cell.number_format if cell.number_format != "General" else ""
                    fmt_str = f" [fmt:{fmt}]" if fmt else ""
                    vals.append(f"[Col{cell.column}({cell.coordinate})] {repr(v)}{fmt_str}")
            if vals:
                print(f"  Row {row_idx}: {' | '.join(vals)}")
            else:
                print(f"  Row {row_idx}: (empty)")
    print()
    wb.close()
